#!/usr/bin/env python3
"""
LinkedIn Export Scraper
Downloads the native XLSX/CSV export from LinkedIn admin analytics.
This is the ONLY reliable way to get ALL posts with full metrics.
"""
import os
import json
import time
import hashlib
from datetime import datetime, date, timedelta
from pathlib import Path

DATA_DIR = Path("data_output")
DOWNLOADS_DIR = DATA_DIR / "linkedin_downloads"
DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
COMPANY_ID = "68624141"
UPDATES_URL = f"https://www.linkedin.com/company/{COMPANY_ID}/admin/analytics/updates/"


def log(m):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] [LI-Export] {m}", flush=True)


def _get_sb():
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not url or not key:
        try:
            import streamlit as st
            try: url = str(st.secrets["SUPABASE_URL"]).strip()
            except: pass
            try: key = str(st.secrets["SUPABASE_SERVICE_KEY"]).strip()
            except: pass
        except: pass
    if not url or not key:
        return None
    from supabase import create_client
    return create_client(url, key)


def export_and_parse():
    """Open LinkedIn admin, set 365 days, click Export, parse XLSX."""
    from playwright.sync_api import sync_playwright

    session_file = DATA_DIR / "linkedin_session_state.json"
    if not session_file.exists():
        log("No session state - run linkedin_login_save.py first")
        return None

    log("Opening LinkedIn admin analytics...")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        context = browser.new_context(
            storage_state=str(session_file),
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            viewport={"width": 1440, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()

        page.goto(UPDATES_URL, timeout=45000, wait_until="domcontentloaded")
        time.sleep(5)

        if "login" in page.url.lower():
            log("LOGIN REQUIRED - run linkedin_login_save.py first")
            browser.close()
            return None

        log(f"Page loaded: {page.url}")

        # Set date to Last 365 days
        try:
            for sel in ["button:has-text('Last')", "button:has-text('days')",
                        "button:has-text('2026')", "button:has-text('2025')"]:
                btns = page.query_selector_all(sel)
                for btn in btns:
                    txt = (btn.inner_text() or "").strip()
                    if any(m in txt for m in ["2025","2026","days","Last","Past"]):
                        btn.click()
                        time.sleep(2)
                        log(f"Clicked date picker: {txt[:40]}")
                        break
                else:
                    continue
                break

            for opt in ["Last 365 days", "Last year", "Past year"]:
                el = page.query_selector(f"button:has-text('{opt}'), li:has-text('{opt}'), [role='option']:has-text('{opt}')")
                if el:
                    el.click()
                    time.sleep(4)
                    log(f"Selected: {opt}")
                    break
        except Exception as e:
            log(f"Date range (non-fatal): {e}")

        time.sleep(3)

        # Scroll to load ALL posts
        log("Scrolling to load all posts...")
        for i in range(50):
            before = page.evaluate("document.querySelectorAll('tr, [role=row]').length")
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(2)
            for btn_sel in ["button:has-text('Show more')", "button:has-text('Load more')"]:
                try:
                    btn = page.query_selector(btn_sel)
                    if btn and btn.is_visible():
                        btn.click()
                        time.sleep(2)
                except Exception:
                    pass
            after = page.evaluate("document.querySelectorAll('tr, [role=row]').length")
            if after == before and i > 5:
                break

        time.sleep(3)

        # Try to click Export button
        download_path = None
        try:
            with page.expect_download(timeout=30000) as dl_info:
                for sel in ["button:has-text('Export')", "[aria-label*='Export' i]",
                            "a:has-text('Export')", "button:has(span:has-text('Export'))"]:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        btn.click()
                        log(f"Clicked Export")
                        break
                time.sleep(3)
                # Confirm if needed
                for opt in ["All updates", "Export", "Download", "Confirm"]:
                    try:
                        confirm = page.query_selector(f"button:has-text('{opt}')")
                        if confirm and confirm.is_visible():
                            confirm.click()
                            time.sleep(2)
                            break
                    except Exception:
                        pass

            download = dl_info.value
            save_path = DOWNLOADS_DIR / f"updates_{date.today()}.xlsx"
            download.save_as(str(save_path))
            download_path = str(save_path)
            log(f"Downloaded: {save_path}")
        except Exception as e:
            log(f"Export download failed: {e}")
            log("Falling back to DOM scraping...")

        # DOM scraping with PAGINATION - click Next to get ALL pages
        posts = []
        page_num = 1
        while True:
            try:
                rows = page.query_selector_all("tr, [role='row']")
                log(f"Page {page_num}: Found {len(rows)} table rows")

            for row in rows:
                try:
                    cells = row.query_selector_all("td, [role='cell']")
                    if len(cells) < 4:
                        # Try getting all text lines
                        txt = row.inner_text().strip()
                        if txt and len(txt) > 20:
                            lines = [l.strip() for l in txt.split("\n") if l.strip()]
                            if len(lines) >= 4:
                                cells_text = lines
                            else:
                                continue
                        else:
                            continue
                    else:
                        cells_text = [c.inner_text().strip() for c in cells]

                    # Skip header rows
                    if any(h in cells_text[0] for h in ["Post title", "Impressions", "Title"]):
                        continue

                    def safe_int(s):
                        try:
                            return int(str(s).replace(",", "").strip() or "0")
                        except:
                            return 0

                    def safe_float(s):
                        try:
                            return float(str(s).replace("%", "").replace(",", "").strip() or "0")
                        except:
                            return 0.0

                    # Extract date from post text if present
                    import re
                    pub_date = None
                    date_match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", cells_text[0])
                    if date_match:
                        pub_date = date_match.group(1)

                    post = {
                        "title":           cells_text[0][:500],
                        "post_type":       cells_text[1] if len(cells_text) > 1 else "",
                        "audience":        cells_text[2] if len(cells_text) > 2 else "",
                        "impressions":     safe_int(cells_text[3]) if len(cells_text) > 3 else 0,
                        "views":           safe_int(cells_text[4]) if len(cells_text) > 4 else 0,
                        "clicks":          safe_int(cells_text[5]) if len(cells_text) > 5 else 0,
                        "ctr":             safe_float(cells_text[6]) if len(cells_text) > 6 else 0,
                        "reactions":       safe_int(cells_text[7]) if len(cells_text) > 7 else 0,
                        "comments":        safe_int(cells_text[8]) if len(cells_text) > 8 else 0,
                        "reposts":         safe_int(cells_text[9]) if len(cells_text) > 9 else 0,
                        "follows":         safe_int(cells_text[10]) if len(cells_text) > 10 else 0,
                        "engagement_rate": safe_float(cells_text[11]) if len(cells_text) > 11 else 0,
                        "published_at":    pub_date,
                    }
                    if post["impressions"] > 0 or post["reactions"] > 0:
                        posts.append(post)
                except Exception:
                    continue
        except Exception as e:
            log(f"DOM scraping error: {e}")

        log(f"DOM scraped: {len(posts)} posts")
        browser.close()

    # Parse XLSX if downloaded
    if download_path:
        xlsx_posts = parse_xlsx(download_path)
        if xlsx_posts:
            posts = xlsx_posts
            log(f"XLSX parsed: {len(posts)} posts (overrides DOM)")

    # Save to Supabase
    if posts:
        save_posts_to_supabase(posts)

    # Save local cache
    cache = {"scraped_at": datetime.utcnow().isoformat(), "posts": posts, "count": len(posts)}
    (DATA_DIR / "linkedin_all_posts.json").write_text(json.dumps(cache, indent=2, default=str))
    log(f"Saved {len(posts)} posts to cache + Supabase")

    return posts


def parse_xlsx(filepath):
    """Parse LinkedIn export XLSX file."""
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.run(["pip", "install", "openpyxl", "-q"])
        import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        posts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h or "").strip() for h in rows[0]]
            for row in rows[1:]:
                if not any(row):
                    continue
                d = dict(zip(headers, row))
                title = str(d.get("Post title", d.get("Title", d.get(headers[0], ""))) or "")
                if not title:
                    continue

                def si(v):
                    try: return int(str(v or "0").replace(",",""))
                    except: return 0

                def sf(v):
                    try: return float(str(v or "0").replace("%","").replace(",",""))
                    except: return 0.0

                posts.append({
                    "title":           title[:500],
                    "post_type":       str(d.get("Post type", d.get("Type", "")) or ""),
                    "audience":        str(d.get("Audience", "") or ""),
                    "impressions":     si(d.get("Impressions", 0)),
                    "views":           si(d.get("Views", 0)),
                    "clicks":          si(d.get("Clicks", 0)),
                    "ctr":             sf(d.get("Click-through rate (CTR)", d.get("CTR", 0))),
                    "reactions":       si(d.get("Reactions", 0)),
                    "comments":        si(d.get("Comments", 0)),
                    "reposts":         si(d.get("Reposts", d.get("Shares", 0))),
                    "follows":         si(d.get("Follows", 0)),
                    "engagement_rate": sf(d.get("Engagement rate", 0)),
                    "published_at":    str(d.get("Post link", d.get("Date", ""))) if d.get("Date") else None,
                })
        log(f"XLSX: {len(posts)} posts from {len(wb.sheetnames)} sheets")
        return posts
    except Exception as e:
        log(f"XLSX parse error: {e}")
        return []


def save_posts_to_supabase(posts):
    """Save all posts to Supabase linkedin_posts table."""
    sb = _get_sb()
    if not sb:
        log("Supabase not configured - skipped")
        return

    rows = []
    for p in posts:
        urn = "export::" + hashlib.sha1(p["title"][:200].encode()).hexdigest()[:16]
        rows.append({
            "urn":             urn,
            "title":           p["title"][:500],
            "post_type":       p.get("post_type", "")[:50],
            "audience":        p.get("audience", "")[:100],
            "published_at":    p.get("published_at"),
            "impressions":     p.get("impressions", 0),
            "views":           p.get("views", 0),
            "clicks":          p.get("clicks", 0),
            "ctr":             p.get("ctr", 0),
            "reactions":       p.get("reactions", 0),
            "comments":        p.get("comments", 0),
            "reposts":         p.get("reposts", 0),
            "follows":         p.get("follows", 0),
            "engagement_rate": p.get("engagement_rate", 0),
            "last_updated":    datetime.utcnow().isoformat(),
        })

    log(f"Upserting {len(rows)} posts to linkedin_posts...")
    errors = 0
    for i in range(0, len(rows), 50):
        try:
            sb.table("linkedin_posts").upsert(rows[i:i+50], on_conflict="urn").execute()
        except Exception as e:
            errors += 1
            if errors <= 3:
                log(f"Upsert error: {e}")
    log(f"Done: {len(rows)} posts, {errors} errors")


if __name__ == "__main__":
    posts = export_and_parse()
    if posts:
        print(f"\nTotal posts captured: {len(posts)}")
        for p in posts[:5]:
            print(f"  {p.get('published_at','?')}: {p.get('title','')[:60]} | {p.get('impressions',0)} imp")
    else:
        print("No posts captured")
